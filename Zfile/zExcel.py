#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :   2020/07/01 10:50
# @Author   :   ZANWEB
# @File     :   zExcel.py in Batch
# @IDE      :   PyCharm

import pandas as pd
import xlrd
import openpyxl
from openpyxl.styles import Color, PatternFill
from pprint import pprint


class ExcelFile:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.sheet = None
        self.book_xlsx = None

    def direct_read_construct(self):
        try:
            self.sheet = []
            file = open(self.excel_path)
            for line in file:
                line_t = line.replace('\n', '').upper()
                list_in_line = line_t.split('\t')
                self.sheet.append(list_in_line)
            file.close()
        except Exception as error:
            print(error)

    def remove_head_tail_direct_construct(self):
        nrows = len(self.sheet)
        start_row = 0
        end_row = 0
        tmp_sheet = []
        for i_index in range(0, nrows):
            tmp_row = []
            if end_row == 0:
                if isinstance(self.sheet[i_index][0], str):
                    if self.sheet[i_index][0][2:5] == '---':
                        if (i_index - start_row) > 3:
                            end_row = i_index
                        else:
                            start_row = i_index
            for j_index in range(0, len(self.sheet[i_index])):
                item = self.sheet[i_index][j_index]
                if isinstance(item, str):
                    item = item.upper()
                    item = item.strip()
                tmp_row.append(item)
            tmp_sheet.append(tmp_row)
        tmp_sheet = tmp_sheet[(start_row + 1):end_row]
        return tmp_sheet

    def pandas_read(self):
        self.sheet = pd.read_excel(self.excel_path, engine='xlrd')

    def xlrd_read(self):
        data = xlrd.open_workbook(self.excel_path)
        self.sheet = data.sheets()[0]

    def openpyxl_read(self):
        self.book_xlsx = openpyxl.load_workbook(self.excel_path)
        self.sheet = self.book_xlsx.worksheets[0]

    def remove_head_tail(self):
        nrows = self.sheet.nrows
        start_row = 0
        end_row = 0
        tmp_sheet = []
        for i_index in range(0, nrows):
            tmp_row = []
            if end_row == 0:
                if isinstance(self.sheet.row_values(i_index)[0], str):
                    if self.sheet.row_values(i_index)[0][2:5] == '---':
                        if (i_index - start_row) > 3:
                            end_row = i_index
                        else:
                            start_row = i_index
            for j_index in range(0, len(self.sheet.row_values(i_index))):
                item = self.sheet.row_values(i_index)[j_index]
                if isinstance(item, str):
                    item = item.upper()
                    item = item.strip()
                tmp_row.append(item)
            tmp_sheet.append(tmp_row)
        tmp_sheet = tmp_sheet[(start_row + 1):end_row]
        return tmp_sheet

    def ie_confirm_change(self):
        selected_items = []
        row_number = self.sheet.max_row
        fill = PatternFill(fill_type='solid', fgColor='00FFFF00')
        for row_index in range(2, row_number + 1):
            cell_item_no_bk_color = self.sheet.cell(row=row_index, column=5).fill.end_color.index
            if cell_item_no_bk_color != '00000000':
                for col in [2, 10, 12]:
                    cell = self.sheet.cell(row=row_index, column=col)
                    cell.fill = fill
                    if col == 2:
                        cell.value = 'Y'
                    elif col == 10:
                        cell.value = 'MO'
                    else:
                        cell.value = 'O'
                tmp_info = []
                tmp_info.append(self.sheet.cell(row=row_index, column=5).value)   # item no
                tmp_info.append(self.sheet.cell(row=row_index, column=6).value)   # item name
                tmp_info.append(self.sheet.cell(row=row_index, column=7).value)   # quantity
                tmp_info.append(self.sheet.cell(row=row_index, column=16).value)  # UnitWeight
                tmp_info.append(self.sheet.cell(row=row_index, column=17).value)  # suffix
                tmp_info.append(self.sheet.cell(row=row_index, column=18).value)  # manifest id
                tmp_info.append(self.sheet.cell(row=row_index, column=19).value)  # version

                selected_items.append(tmp_info)
        return selected_items

    def save_xlsx(self):
        self.book_xlsx.save(self.excel_path)


def test01():
    ex_file = ExcelFile('C:\\Users\\zzimo\\Desktop\\构件清单\\1\\20-2000793001-构件组成清单.xls')
    ex_file.xlrd_read()
    sheet_data = ex_file.remove_head_tail()
    pprint(sheet_data)


def test02():
    ex_file = ExcelFile('C:\\Users\\zzimo\\Desktop\\构件清单\\1\\2000793001-2211-001C IE.xlsx')
    ex_file.openpyxl_read()
    items = ex_file.ie_confirm_change()
    ex_file.save_xlsx()
    pprint(items)


if __name__ == '__main__':
    test02()

